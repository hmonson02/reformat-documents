import openpyxl
print(openpyxl.__version__)
import os

def semiColonAdder(filename, directory, column, sheetname, newfilename): #removes spaces from authour column, makes semicolon seperated
    os.chdir(directory) #DIRECTORY
    wb = openpyxl.load_workbook(filename) #PUT IN FILE NAME. TAKES XLSX, XLTM OR XLSM
    print(wb.sheetnames)
    sheet = wb[sheetname]
    collength = len(sheet[column]) #COLUMN LETTER WITH AUTHOURS
    lst = []
    strn = ''
    for i in range(3,collength + 1):
        val = sheet.cell(row = i, column = 11).value
        if val == None:
            break
        else:
            lst = val.split('\n\n')
            for j in range(len(lst)):
                lst[j] = lst[j] + ';'
            strn = ''.join(lst)
            sheet.cell(row = i, column = 11).value = strn
            print(sheet.cell(row = i, column = 11).value)
            lst = []
        wb.save(newfilename) #NEW FILE NAME

#semiColonAdder('newsample.xlsm', '/Users/hayleymonson/Desktop/Bib analysis', 'K', 'newsample2.xlsm', 'Sheet 1')

def spaceRemover(filename, directory, column, sheetname, newfilename): #removes random numbers and date after SO
    os.chdir(directory) #DIRECTORY
    wb = openpyxl.load_workbook(filename) #PUT IN FILE NAME. TAKES XLSX, XLTM OR XLSM
    print(wb.sheetnames)
    sheet = wb[sheetname]
    collength2 = len(sheet[column])
    lst = []
    for i in range(3, collength2 + 1):
        val = sheet.cell(row = i, column = 7).value
        if val == None:
            break
        else:
            print(val)
            print(type(val))
            lst = val.split('.', 1)
            sheet.cell(row = i, column = 7).value = lst[0]
            print(sheet.cell(row = i, column = 7).value)
            lst = []
        wb.save(newfilename)
    
#spaceRemover('Book1.xlsm', '/Users/hayleymonson/Desktop/Bib analysis', 'G', 'Sheet1', 'Book2.xlsx')
 
def institutionFixer(filename, directory, column, sheetname, newfilename): #semicolon seperates institutions
    os.chdir(directory) #DIRECTORY
    wb = openpyxl.load_workbook(filename) #PUT IN FILE NAME. TAKES XLSX, XLTM OR XLSM
    print(wb.sheetnames)
    sheet = wb[sheetname]
    collength = len(sheet[column]) #COLUMN LETTER WITH AUTHOURS
    lst = []
    strn = ''
    for i in range(3, collength + 1):
        val = sheet.cell(row = i, column = 15).value
        if val == None:
            break
        else:
            lst = val.split('\n\n')
            for j in range(len(lst)):
                lst[j] = lst[j][:-1] + ','
                print(lst[j])
            strn = ''.join(lst)
            sheet.cell(row = i, column = 15).value = strn
            print(sheet.cell(row = i, column = 15).value)
            lst = []
        wb.save(newfilename)
        
        

semiColonAdder('shortname.xlsm', '/Users/hayleymonson/Desktop/Victoria', 'K', 'Sheet1', 'sampleedited.xls')
#spaceRemover('covid.xlsm', '/Users/hayleymonson/Desktop/Bib analysis', 'G', 'Sheet1', 'covidmod2.xls')  
#institutionFixer('covid.xlsm', '/Users/hayleymonson/Desktop/Bib analysis', 'O', 'Sheet1', 'covidmod3.xls')    