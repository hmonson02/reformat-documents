import openpyxl
import os
import string
import itertools
from openpyxl import Workbook
import re

def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)
    sheet=workbook.active


def mainFunc(filename, directory, sheetname, newfilename):
    os.chdir(directory)
    wb = openpyxl.load_workbook(filename)
    sheet2=wb[sheetname]
    workbook = Workbook()
    sheet=workbook.active
    alphabet_string=string.ascii_uppercase
    alphabet_list=list(alphabet_string)
    #create_workbook(newfilename)
    headerList=['Authors',	'Author(s) ID',	'Title','Year',	'Source title',	'Volume',	'Issue',	'Art. No.',	'Page start',	'Page end',	'Page count',	'Cited by',	'DOI',	'Link',	'Affiliations',	'Authors with affiliations',	'Abstract',	'Author Keywords',	'Index Keywords',	'References',	'Correspondence Address',	'Editors',	'Publisher', 'ISSN','ISBN', 'CODEN', 'PubMed ID', 'Language of Original Document', 'Abbreviated Source Title', 'Document Type', 'Publication Stage',	'Open Access','Source', 'EID']
    alphalist1=[None]*26
    alphalist2=[None]*8
    for i in range(len(alphabet_list)):
        alphalist1[i]=alphabet_list[i] + '1'
        
    for i in range(len(alphalist2)):
        alphalist2[i]='A'+ alphabet_list[i] + '1'
    
    fullalphalist=alphalist1 + alphalist2
    print(fullalphalist)
    
        
    for i in range(len(headerList)):
        sheet[fullalphalist[i]]=headerList[i]
        
    mr = sheet2.max_row
    mc = sheet2.max_column
    
    for i in range(3, mr):
        c=sheet2.cell(row = i, column = 6)
        sheet.cell(row=i, column = 3).value=c.value #title
        
        c=sheet2.cell(row = i, column = 3)
        sheet.cell(row=i, column = 33).value=c.value #database/source
        
        c=sheet2.cell(row = i, column = 5)
        sheet.cell(row=i, column = 31).value=c.value #article status
        
        c=sheet2.cell(row = i, column = 14)
        sheet.cell(row=i, column = 23).value=c.value #publisher
        
        c=sheet2.cell(row = i, column = 16)
        sheet.cell(row=i, column = 14).value=c.value #link
        
        c=sheet2.cell(row = i, column = 20)
        sheet.cell(row=i, column = 17).value=c.value #abstract
        
        c=sheet2.cell(row = i, column = 21)
        sheet.cell(row=i, column = 12).value=c.value #cited ref count
        
        c=sheet2.cell(row = i, column = 23)
        sheet.cell(row=i, column = 13).value=c.value #DOI
        
        c=sheet2.cell(row = i, column = 24)
        sheet.cell(row=i, column = 28).value=c.value #language
        
        c=sheet2.cell(row = i, column = 28)
        sheet.cell(row=i, column = 30).value=c.value #publication type
        
        c=sheet2.cell(row = i, column = 32)
        sheet.cell(row=i, column = 4).value=c.value #year
        
        c=sheet2.cell(row = i, column = 12)
        sheet.cell(row=i, column = 21).value=c.value #correspondence address
        
    #loop through authour column and remove enters and add commas, then replace the corresponding cell in the first column of the new spread sheet with the authour list.
    
    
    for i in range(3,len(sheet2['H'])):
        val = sheet2.cell(row = i,column = 8).value
        if val != None:
            lst = val.split('\n\n')
            for j in range(len(lst)):
                lst[j]=lst[j]+','
            newval=''.join(lst)
            sheet.cell(row=i,column=1).value=newval
    
    #removes crap from after institutions
    for i in range(3,len(sheet2['G'])):
        val = sheet2.cell(row = i,column = 7).value
        if val != None:
            lst=val.split('.',1)
            print(lst)
            sheet.cell(row=i,column=5).value=lst[0]
            print(sheet.cell(row=i,column=5).value)
    
    #semi-colon seperates keywords
    for i in range(3,len(sheet2['O'])):
        val = sheet2.cell(row = i,column = 15).value
        if val != None:
            lst = val.split('\n\n')
            for j in range(len(lst)):
                lst[j]=lst[j]+';'
            newval=''.join(lst)
            sheet.cell(row=i,column=18).value=newval
    
    
    #references, authours with affiliations, affiliations
    
    for i in range(3, len(sheet['K'])):
        val = sheet2.cell(row = i,column = 11).value
        if val != None:
            val = re.sub('(\n[(][A-Za-z]*[)])', ';', val)
            val = re.sub('($[(][[A-Za-z]*[)])', '', val)
            #val = re.sub(('(\n[(][[A-Za-z]*[,][ ]]*[)])', ';', val)
            print(val)
            sheet.cell(row=i, column = 16).value=val
            
        
    workbook.save(newfilename)
        
        

    
mainFunc('embase_sample.xlsm', '/Users/hayleymonson/Desktop/Victoria', 'citations', 'alteredsample.xls')

#authour in H


    

#affiliations: department, location, city, country; next entry
#authours with affiliations: name, initial., same as affiliations

    

